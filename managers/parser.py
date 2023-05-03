from abc import ABC
import openpyxl as pyxl


class DocumentParser(ABC):
    def __init__(self, table_name: str) -> None:
        """
        Class Constructor.

        Args:
        table_name (str): The name of the .xlsx file to be loaded.
        """
        self.workbook = pyxl.load_workbook(table_name, read_only=True)
        self.sheet = self.workbook.active
        self._data = []


class FileReader(DocumentParser):
    def __init__(self, table_name: str) -> None:
        """
        Class Constructor.

        Args:
        table_name (str): The name of the .xlsx file to be loaded.
        """
        super().__init__(table_name)
        self.values_only = True
        self.min_col = 2
        self.min_row = 4
        self.max_row = 3
        self.rows = []
        self.rows_len = {}

    def read_fields_values(self) -> None:
        """Read cell values from the sheet and stores them in the _data list"""
        for row in self.sheet.iter_rows(
            values_only=self.values_only,
            min_col=self.min_col,
            min_row=self.min_row
        ):
            self._data.append(row)

    def read_column_names(self) -> list:
        """Read the column names from the sheet"""
        next_key = None

        for row in self.sheet.iter_rows(
            min_col=self.min_col,
            min_row=self.min_row - 3,
            max_row=self.max_row,
        ):
            for cell in row:
                if cell.value:
                    next_key = cell.value
                    self.rows_len[next_key] = 1
                elif next_key:
                    self.rows_len[next_key] += 1
            next_key = None

            self.rows.append([r.value for r in row if r.value])

    def get_crossed_names(self, exclude_fields: list) -> list:
        """
        Get the crossed fields from the rows list
        
        Args:
        exclude_fields (list): List of columns to be excluded

        Returns:
        A list of crossed columns
        """
        try:
            crossed_names = set()
            for a in self.rows[0]:
                if a not in exclude_fields:
                    for b in self.rows[1][:self.rows_len[a]]:
                        self.rows_len[a] -= 1
                        for c in self.rows[2][:self.rows_len[a]]:
                            crossed_names.add(f"{a}_{b}_{c}")
            return crossed_names
        except IndexError:
            raise ("Incorrect table view")

    def get_all_col_names(self, exclude_fields: list) -> list:
        """
        Get all column names

        Args:
        exclude_fields (list): List of columns to be excluded

        Returns:
        A list of column names.
        """
        crossed_names = self.get_crossed_names(exclude_fields)
        common_names = self.rows[0]
        filtered_names = set()

        for name in common_names:
            if not any(name_.startswith(name) for name_ in crossed_names):
                filtered_names.add(name)

        return sorted(crossed_names | filtered_names)

    @property
    def data(self) -> list:
        """
        Returns _data

        Returns:
        A list with the sheet data
        """
        return self._data
